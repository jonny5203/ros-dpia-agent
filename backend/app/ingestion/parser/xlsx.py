from __future__ import annotations
import io
from app.ingestion.parser.base import Parser
from app.ingestion.types import ParsedDocument, ParsedSection


class XlsxParser(Parser):
    name = "openpyxl"
    version = "openpyxl>=3.1"

    def parse(self, data: bytes, filename: str) -> ParsedDocument:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sections: list[ParsedSection] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    sections.append(ParsedSection(
                        text=" | ".join(cells),
                        section_title=sheet.title,
                    ))
        return ParsedDocument(sections=sections, parser=self.name,
                              parser_version=self.version)
